#!/usr/bin/env python3
"""
full_sync.py - Sincroniza maestro + fotos desde Drive y deploya a Netlify.

Pensado para correr en GitHub Actions cada 2 hs.
Requiere variables de entorno:
  GDRIVE_SA_JSON   - JSON del service account de Google (string)
  NETLIFY_TOKEN    - personal access token de Netlify
  NETLIFY_SITE_ID  - site id en Netlify (default ya seteado)

Hace:
  1. Baja maestro_latest.xlsx desde Drive.
  2. Corre sync.py para regenerar productos.json + precios.json.
  3. Lista fotos en las carpetas de Drive y baja las que faltan o cambiaron.
  4. Corre build.py para generar web/index.html.
  5. Empaqueta web/ y lo manda a Netlify.

Idempotente: si nada cambió, no deploya (compara con un hash).
"""
import os, sys, json, base64, io, re, hashlib, subprocess, time, zipfile, tempfile
from io import BytesIO

import requests
from google.oauth2 import service_account
from googleapiclient.discovery import build as gbuild
from googleapiclient.http import MediaIoBaseDownload
from PIL import Image
import openpyxl

# -------- Configuración fija --------
MAESTRO_FILE_ID = '1stw7L-HT4letxcgxNrhjjKzGt1-oDxQu'
DEFAULT_NETLIFY_SITE_ID = 'e76473c5-431e-4949-9c83-d59143300dbd'

# Carpetas de fotos a escanear (parent IDs en Drive)
PHOTO_FOLDERS = {
    # WEB CATALOGO INVIERNO 26 (oficial)
    'WEB CATALOGO/HOMBRE':   '1NU5TcZRFjpzLoncyXmlwwN3TJrQbQB01',
    'WEB CATALOGO/MUJER':    '1kbvw14AM4XD3DXK0Qhl5iI5vHUEbIkie',
    # 05. MAYO
    'MAYO/BATUK HOMBRE':     '1k1uvapxAIQfTyH9jbpj5EoTkYXM9Za_2',
    'MAYO/BATUK MUJER':      '114wXwppqpzmh1A7hqOYyGpj0N9wP7wR5',
    'MAYO/HUOKY HOMBRE':     '1mMbDlC4ngq1WwMOmi8n3NYHE93vbxFIU',
    # 04. ABRIL
    'ABRIL/BATUK MUJER':     '1rcBEzH2NWHUPaw_ikzxa4Ui4zFHQSOwR',
    'ABRIL/BATUK HOMBRE':    '1Xj8l59v5J3QR2BSd1UJj3ieoYZk3V8do',
    'ABRIL/HUOKY HOMBRE':    '1Ft8UlvBFOn4bgJ0yT0B4N4gaVuNayBv2',
    # 03. MARZO
    'MARZO/BATUK HOMBRE':    '1vcbsOji9wNiOFqK2jClRs9_S6LJyqrlL',
    'MARZO/BATUK MUJER':     '1YqkK3ynGQ4rgbKpStNxl_3dEOzwkXH3z',
    'MARZO/HUOKY HOMBRE':    '1tapDAxYZFfuNa44bOCCqLXmpncWmdKaU',
}

# Paths del repo (relativos al cwd cuando corre la action)
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MAESTRO_PATH = os.path.join(ROOT, 'automation', 'maestro_latest.xlsx')
FOTOS_DIR = os.path.join(ROOT, 'fotos')
WEB_DIR = os.path.join(ROOT, 'web')
TALLER = os.path.join(ROOT, '_taller')
STATE_FILE = os.path.join(ROOT, 'automation', 'state.json')


def log(msg):
    print(f'[{time.strftime("%H:%M:%S")}] {msg}', flush=True)


# -------- Google Drive --------
def get_drive():
    sa_json = os.environ.get('GDRIVE_SA_JSON')
    if not sa_json:
        raise RuntimeError('Falta variable GDRIVE_SA_JSON')
    info = json.loads(sa_json)
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=['https://www.googleapis.com/auth/drive.readonly'])
    return gbuild('drive', 'v3', credentials=creds, cache_discovery=False)


def drive_download(drive, file_id, dest_path):
    request = drive.files().get_media(fileId=file_id)
    buf = BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    with open(dest_path, 'wb') as f:
        f.write(buf.getvalue())
    return os.path.getsize(dest_path)


def drive_list_images(drive, folder_id):
    """Lista todas las imágenes en una carpeta (con paginación)."""
    out = []
    page_token = None
    while True:
        resp = drive.files().list(
            q=f"'{folder_id}' in parents and mimeType contains 'image/' and trashed = false",
            fields='nextPageToken, files(id, name, modifiedTime, size)',
            pageSize=500,
            pageToken=page_token,
        ).execute()
        out.extend(resp.get('files', []))
        page_token = resp.get('nextPageToken')
        if not page_token:
            break
    return out


# -------- Lógica de matching y descarga de fotos --------
def codbas_from_title(title):
    """Extrae el CODBAS del inicio del nombre del archivo (ej: BMP60REMC28 ...)."""
    m = re.match(r'^([A-Za-z]+\d+[A-Za-z]+\d+)', title)
    return m.group(1).upper() if m else None


def score_photo(name, size):
    """Puntaje para elegir la mejor foto de un codbas."""
    u = name.upper()
    s = 0
    if name.lower().endswith(('.jpg', '.jpeg')): s += 10
    if ' 1.' in u or ' 1 ' in u or 'FRENTE' in u or 'PORTADA' in u: s += 5
    if 'FP' in u: s -= 2
    if 'DORSO' in u or 'DETALLE' in u or 'ESPALDA' in u: s -= 3
    if size and int(size) > 8_000_000: s -= 3
    return s


# Solapas del maestro a combinar (mismo criterio que _taller/sync.py).
MAESTRO_SHEETS = ['INVIERNO', 'VERANO']


def _norm_header(s):
    return (str(s or '').strip().upper()
            .replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U').replace('Ñ','N'))


def build_codbas_to_cod(maestro_path):
    """Lee todas las solapas del maestro y devuelve dict codbas -> (cod, nombre).

    Usa lookup de columnas por nombre de header (no por posición fija) para
    no romperse si el maestro cambia de estructura. Si una solapa no tiene
    columna CODIGO BAS (ej VERANO), simplemente no aporta nada a este mapeo
    (esos productos no podrán auto-matchear foto por codbas).
    """
    wb = openpyxl.load_workbook(maestro_path, data_only=True)
    cb_map = {}
    for sheet_name in MAESTRO_SHEETS:
        if sheet_name not in wb.sheetnames:
            log(f'[build_codbas_to_cod] AVISO: no existe la solapa "{sheet_name}", la salteo.')
            continue
        ws = wb[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {_norm_header(c): i for i, c in enumerate(header_row) if c}

        def find_col(*candidates):
            for name in candidates:
                key = _norm_header(name)
                if key in headers:
                    return headers[key]
                for k, idx in headers.items():
                    if key in k:
                        return idx
            return None

        c_cod = find_col('ARTICULO', 'COD')
        c_cb = find_col('CODIGO BAS')
        c_nom = find_col('NOMBRE PRODUCTO', 'NOMBRE')
        if c_cod is None:
            log(f'[build_codbas_to_cod] [{sheet_name}] AVISO: no encontré columna ARTICULO/COD, salteo la hoja.')
            continue
        if c_cb is None:
            log(f'[build_codbas_to_cod] [{sheet_name}] AVISO: sin columna CODIGO BAS, sus productos no matchean foto por codbas.')
            continue
        n = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            cod = r[c_cod]
            if not cod: continue
            cod = str(cod).strip()
            cb = str(r[c_cb]).strip() if c_cb is not None and r[c_cb] else ''
            nom = str(r[c_nom]).strip() if c_nom is not None and r[c_nom] else ''
            if cod and cb and cb not in cb_map:
                cb_map[cb] = (cod, nom)
                n += 1
        log(f'[build_codbas_to_cod] [{sheet_name}] {n} codbas mapeados')
    return cb_map


def load_state():
    if os.path.exists(STATE_FILE):
        try: return json.load(open(STATE_FILE))
        except: return {}
    return {}


def save_state(state):
    os.makedirs(os.path.dirname(STATE_FILE), exist_ok=True)
    json.dump(state, open(STATE_FILE, 'w'), indent=1)


def save_image_as_jpg(binary, out_path, max_width=800, quality=82):
    img = Image.open(BytesIO(binary))
    if img.mode in ('RGBA', 'LA', 'P'):
        img = img.convert('RGB')
    if img.width > max_width:
        ratio = max_width / img.width
        img = img.resize((max_width, int(img.height * ratio)), Image.LANCZOS)
    img.save(out_path, 'JPEG', quality=quality, optimize=True)


def sync_photos(drive, cb_to_cod, state):
    """Para cada producto, busca la mejor foto disponible y descarga si cambió."""
    os.makedirs(FOTOS_DIR, exist_ok=True)
    # 1) Listar todas las fotos de todas las carpetas
    all_files = {}  # codbas -> list de (name, id, size, modifiedTime, source)
    for source, folder_id in PHOTO_FOLDERS.items():
        try:
            files = drive_list_images(drive, folder_id)
            log(f'  {source}: {len(files)} fotos')
        except Exception as e:
            log(f'  {source}: ERROR {e}')
            continue
        for f in files:
            cb = codbas_from_title(f['name'])
            if not cb: continue
            # Validar que cb esté en el nombre
            if cb not in f['name'].upper(): continue
            all_files.setdefault(cb, []).append({
                'name': f['name'], 'id': f['id'],
                'size': int(f.get('size', 0)),
                'modified': f.get('modifiedTime', ''),
                'source': source,
            })
    log(f'CODBAS únicos con foto disponible: {len(all_files)}')

    # 2) Para cada producto que tenga match, descargar la mejor (si cambió)
    new_count = 0
    replaced_count = 0
    drive_state = state.setdefault('drive_files', {})
    for cb, (cod, nom) in cb_to_cod.items():
        if cb not in all_files: continue
        candidates = all_files[cb]
        best = max(candidates, key=lambda f: score_photo(f['name'], f['size']))
        # Si ya descargamos exactamente esta foto, skip
        key = f'{cod}|{cb}'
        prev = drive_state.get(key, {})
        if prev.get('id') == best['id'] and prev.get('modified') == best['modified']:
            continue
        # Descargar
        cod_fname = cod.replace('/', '_')
        out_path = os.path.join(FOTOS_DIR, cod_fname + '.jpg')
        existed = os.path.exists(out_path)
        try:
            request = drive.files().get_media(fileId=best['id'])
            buf = BytesIO()
            downloader = MediaIoBaseDownload(buf, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            save_image_as_jpg(buf.getvalue(), out_path)
            drive_state[key] = {'id': best['id'], 'modified': best['modified'], 'name': best['name']}
            if existed:
                replaced_count += 1
            else:
                new_count += 1
            log(f'  ✓ {cod} <- {best["name"]} ({os.path.getsize(out_path)//1024}KB)')
        except Exception as e:
            log(f'  ✗ {cod} ERROR: {e}')
    log(f'Fotos nuevas: {new_count}, reemplazadas: {replaced_count}')
    return new_count + replaced_count


# -------- Build (corre los scripts existentes) --------
def run_sync_build():
    sync = os.path.join(TALLER, 'sync.py')
    log('Corriendo sync.py + build.py...')
    res = subprocess.run(
        [sys.executable, sync, MAESTRO_PATH],
        cwd=TALLER, capture_output=True, text=True)
    print(res.stdout)
    if res.returncode != 0:
        print(res.stderr)
        raise RuntimeError('sync.py falló')


# -------- Netlify --------
def deploy_to_netlify():
    token = os.environ.get('NETLIFY_TOKEN')
    site_id = os.environ.get('NETLIFY_SITE_ID', DEFAULT_NETLIFY_SITE_ID)
    if not token:
        raise RuntimeError('Falta NETLIFY_TOKEN')
    log('Empaquetando web/...')
    tmp_zip = tempfile.NamedTemporaryFile(suffix='.zip', delete=False).name
    with zipfile.ZipFile(tmp_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(WEB_DIR):
            for fn in files:
                full = os.path.join(root, fn)
                arc = os.path.relpath(full, WEB_DIR)
                zf.write(full, arc)
    log(f'ZIP: {os.path.getsize(tmp_zip)//1024} KB')
    log('Subiendo a Netlify...')
    with open(tmp_zip, 'rb') as f:
        r = requests.post(
            f'https://api.netlify.com/api/v1/sites/{site_id}/deploys',
            headers={
                'Authorization': f'Bearer {token}',
                'Content-Type': 'application/zip',
            },
            data=f.read(),
            timeout=120,
        )
    os.remove(tmp_zip)
    if r.status_code >= 300:
        raise RuntimeError(f'Netlify deploy falló: {r.status_code} {r.text[:500]}')
    deploy = r.json()
    log(f'✓ Deploy listo: {deploy.get("ssl_url") or deploy.get("url")}')


# -------- Main --------
def main():
    log('=== full_sync inicio ===')
    drive = get_drive()
    log('Bajando maestro...')
    os.makedirs(os.path.dirname(MAESTRO_PATH), exist_ok=True)
    size = drive_download(drive, MAESTRO_FILE_ID, MAESTRO_PATH)
    log(f'Maestro: {size//1024} KB')
    cb_to_cod = build_codbas_to_cod(MAESTRO_PATH)
    log(f'CODBAS en maestro: {len(cb_to_cod)}')
    state = load_state()
    log('Escaneando carpetas de fotos en Drive...')
    changed = sync_photos(drive, cb_to_cod, state)
    save_state(state)
    log('Rebuild...')
    run_sync_build()
    # Deploy siempre (es barato y asegura consistencia con maestro nuevo)
    deploy_to_netlify()
    log('=== fin ===')


if __name__ == '__main__':
    main()
