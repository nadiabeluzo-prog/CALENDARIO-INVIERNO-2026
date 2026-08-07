#!/usr/bin/env python3
"""
Re-extrae productos.json y precios.json desde un xlsx del maestro
(de Drive o local) y dispara build.py para regenerar el calendario.

Uso:
  python3 sync.py /ruta/al/maestro.xlsx
"""
import json, os, sys, subprocess, re

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TALLER = os.path.join(ROOT, '_taller')
PROD_OUT = os.path.join(TALLER, 'productos.json')
PRECIOS_OUT = os.path.join(TALLER, 'precios.json')
BUILD = os.path.join(TALLER, 'build.py')

# rubro -> prefix usado en las URLs de batuk.com.ar
RUBRO_PREFIX = {
    'BABUCHA': 'Pantalon', 'BLUSA': 'Producto', 'BODY': 'Body',
    'BUZO': 'Buzo', 'BUZO C/CAPUCHA': 'Buzo', 'BUZO MEDIO CIERRE': 'Buzo',
    'BUZO S/CAPUCHA': 'Buzo', 'CAMISA': 'Producto', 'CAMPERA': 'Campera',
    'CAMPERA DENIM': 'Jean', 'CARGO': 'Pantalon', 'CHINO': 'Producto',
    'CHOMBA': 'Producto', 'CREW': 'Buzo', 'DENIM': 'Jean',
    'FALDA': 'Falda', 'HOOD': 'Buzo', 'HOODIE': 'Buzo',
    'LEGGING': 'Producto', 'MUSCULOSA': 'Producto', 'PANTALON': 'Pantalon',
    'REMERA': 'Remera', 'REMERA ML': 'Remera', 'REMERA SIN MANGAS': 'Remera',
    'SHORT': 'Short', 'VESTIDO': 'Vestido', 'ZIPHOOD': 'Buzo',
}


def parse_money(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s == '-':
        return None
    s = s.replace('$', '').replace(' ', '').replace('\xa0', '')
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def title_es(s):
    s = (s or '').strip().lower()
    return '-'.join(w.capitalize() for w in re.split(r'\s+', s) if w)


# Solapas del maestro a combinar en un solo calendario. Cada temporada puede tener
# sus propias columnas (ej VERANO no tiene RUBRO/COLOR BAS), así que se detectan
# por separado para cada hoja.
SHEETS = ['INVIERNO', 'VERANO']


def _norm(s):
    return (str(s or '').strip().upper()
            .replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U').replace('Ñ','N'))


def process_sheet(ws, sheet_name, by_cod):
    """Lee una hoja del maestro y vuelca sus productos en el dict compartido by_cod."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {_norm(c): i for i, c in enumerate(header_row) if c}

    def find_col(*candidates, required=True):
        for name in candidates:
            key = _norm(name)
            if key in headers:
                return headers[key]
            for k, idx in headers.items():
                if key in k:
                    return idx
        if required:
            raise KeyError(f'[{sheet_name}] Columna no encontrada: {candidates}. Headers disponibles: {list(headers)}')
        return None

    C_ORIGEN  = find_col('ORIGEN')
    C_MARCA   = find_col('MARCA')
    # IMPORTANTE: 'ARTICULO' es el código corto (ej B-1229) que coincide con los
    # nombres de archivo en /fotos y con las URLs de batuk.com.ar. 'CODIGO BAS' es
    # un código interno distinto (ej BHP50DENN07). Buscar ARTICULO primero para
    # no matchear por error contra CODIGO BAS (que contiene la substring "COD").
    C_COD     = find_col('ARTICULO', 'COD')
    C_CAT     = find_col('CATEGORIA PLAN','CATEGORIA', required=False)
    # RUBRO y COLOR BAS quedaron opcionales: la solapa VERANO no las tiene todavía.
    C_RUBRO   = find_col('RUBRO', required=False)
    C_BOTONERA= find_col('BOTONERA')
    C_DESC    = find_col('DESCRIPCION A PRODUCCION','DESCRIPCION', required=False) or 9
    C_MES_ING = find_col('MES INGRESO', required=False) or 10
    C_MES     = find_col('MES INGRESO PRODUCCION')
    C_SEMANA  = find_col('SEMANA DEL MES','SEMANA')
    C_STATUS  = find_col('STATUS')
    C_NOMBRE  = find_col('NOMBRE PRODUCTO','NOMBRE')
    C_COLOR_BAS = find_col('COLOR BAS','COLOR', required=False)
    C_UNID    = find_col('UNID PEDIDAS','UNID','UNIDADES PEDIDAS','UNIDADES')
    C_CORTADO = find_col('CORTADO','UNID CORTADAS')
    # COSTO OB ya no vive en esta hoja (se movió a "05-COSTOS OBJETIVOS"). La dejamos
    # opcional para que sync.py no crashee; si no está, no se calculan precios.
    C_COSTO_OB= find_col('COSTO OB','COSTO OBJETIVO','COSTO', required=False)
    print(f'[sync.py] [{sheet_name}] Columnas: ORIGEN={C_ORIGEN}, MARCA={C_MARCA}, COD={C_COD}, CAT={C_CAT}, RUBRO={C_RUBRO}, COLOR={C_COLOR_BAS}, NOMBRE={C_NOMBRE}, MES={C_MES}, STATUS={C_STATUS}, UNID={C_UNID}, CORTADO={C_CORTADO}, COSTO={C_COSTO_OB}')
    if C_CAT is None:
        print(f'[sync.py] [{sheet_name}] AVISO: no se encontró columna CATEGORIA. Categoría quedará vacía.')
    if C_RUBRO is None:
        print(f'[sync.py] [{sheet_name}] AVISO: no se encontró columna RUBRO. Rubro quedará vacío.')
    if C_COLOR_BAS is None:
        print(f'[sync.py] [{sheet_name}] AVISO: no se encontró columna COLOR BAS. Color quedará vacío.')
    if C_COSTO_OB is None:
        print(f'[sync.py] [{sheet_name}] AVISO: no se encontró columna COSTO. No se calcularán precios para esta hoja.')

    n_rows = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        cod = r[C_COD]
        if not cod:
            continue
        cod = str(cod).strip()
        if not cod:
            continue
        marca = (r[C_MARCA] or '').strip() if r[C_MARCA] else ''
        if marca.upper() == 'RS':
            continue
        def _s(v):
            if v is None: return ''
            return str(v).strip()
        if cod not in by_cod:
            by_cod[cod] = {
                'COD': cod,
                'NOMBRE': _s(r[C_NOMBRE]),
                'MES INGRESO PRODUCCION': _s(r[C_MES]),
                'cantidad': 0,
                'cortado': 0,
                'rubro': _s(r[C_RUBRO]) if C_RUBRO is not None else '',
                '_colors': set(),
                'categoria': _s(r[C_CAT]) if C_CAT is not None else '',
                'marca': marca,
                'status': _s(r[C_STATUS]),
                'origen': _s(r[C_ORIGEN]),
                'botonera': _s(r[C_BOTONERA]),
                'semana': _s(r[C_SEMANA]),
                'temporada': sheet_name,
                '_costo': None,
            }
        n_rows += 1
        c = r[C_CORTADO]
        if isinstance(c, (int, float)):
            by_cod[cod]['cortado'] += int(c)
        u = r[C_UNID]
        if isinstance(u, (int, float)):
            by_cod[cod]['cantidad'] += int(u)
        if C_COLOR_BAS is not None:
            col = (r[C_COLOR_BAS] or '').strip() if r[C_COLOR_BAS] else ''
            if col:
                by_cod[cod]['_colors'].add(col.upper())
        if C_COSTO_OB is not None and by_cod[cod]['_costo'] is None:
            cnum = parse_money(r[C_COSTO_OB])
            if cnum is not None:
                by_cod[cod]['_costo'] = cnum
    print(f'[sync.py] [{sheet_name}] {n_rows} filas procesadas')


def main(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    by_cod = {}
    any_sheet = False
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f'[sync.py] AVISO: no se encontró la solapa "{sheet_name}" en el maestro. La salteo.')
            continue
        process_sheet(wb[sheet_name], sheet_name, by_cod)
        any_sheet = True
    if not any_sheet:
        raise RuntimeError(f'No se encontró ninguna de las solapas esperadas {SHEETS} en el maestro. '
                            f'Solapas disponibles: {wb.sheetnames}')

    productos = []
    precios = {}
    for cod, p in by_cod.items():
        rubro_u = p['rubro'].strip().upper()
        prefix = RUBRO_PREFIX.get(rubro_u, 'Producto')
        slug = f"{prefix}-{title_es(p['NOMBRE'])}"
        url_prod = f"https://www.batuk.com.ar/productos/{slug}/"
        url_busc = (
            f"https://www.google.com/search?tbm=isch&q="
            f"batuk+{p['NOMBRE'].lower().replace(' ', '+')}+site%3Abatuk.com.ar"
        )
        productos.append({
            'MES INGRESO PRODUCCION': p['MES INGRESO PRODUCCION'],
            'NOMBRE': p['NOMBRE'],
            'COD': cod,
            'cantidad': p['cantidad'],
            'cortado': p['cortado'],
            'rubro': p['rubro'],
            'color': ', '.join(sorted(p['_colors'])),
            'categoria': p['categoria'],
            'marca': p['marca'],
            'status': p['status'],
            'origen': p['origen'],
            'botonera': p['botonera'],
            'semana': p['semana'],
            'temporada': p.get('temporada', ''),
            'web_prefix': prefix,
            'slug': slug,
            'url_producto': url_prod,
            'url_buscar': url_busc,
        })
        if p['_costo'] is not None:
            costo = p['_costo']
            # Importados: el costo en el maestro suele venir tipeado como "41.828"
            # interpretando el punto como separador de miles (formato AR), pero
            # openpyxl lo lee como decimal (41,828). Si origen=IMPORTADO y el
            # costo es absurdamente chico para una prenda, multiplico x1000.
            if (p['origen'] or '').strip().upper() == 'IMPORTADO' and costo < 1000:
                costo = costo * 1000
            precios[cod] = round(costo * 2.16)

    with open(PROD_OUT, 'w', encoding='utf-8') as f:
        json.dump(productos, f, ensure_ascii=False, indent=1)
    with open(PRECIOS_OUT, 'w', encoding='utf-8') as f:
        json.dump(precios, f, ensure_ascii=False)

    print('productos.json:', len(productos), 'filas')
    print('precios.json:  ', len(precios), 'CODs con precio')
    print('--- build ---')
    subprocess.run(['python3', BUILD], check=True)


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('Uso: python3 sync.py /ruta/al/maestro.xlsx', file=sys.stderr)
        sys.exit(1)
    main(sys.argv[1])
